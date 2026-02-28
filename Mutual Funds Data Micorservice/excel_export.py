"""
Excel Export Module
====================
Saves mutual fund NAV data to Excel files using openpyxl/pandas.
"""

import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from config import EXCEL_OUTPUT_DIR

logger = logging.getLogger(__name__)


def save_to_excel(df: pd.DataFrame, filename: str = None) -> str:
    """
    Save DataFrame to a formatted Excel file.

    Args:
        df: DataFrame containing fund data
        filename: Optional filename (auto-generated if not provided)

    Returns:
        Path to the saved Excel file
    """
    os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"mutual_funds_nav_{timestamp}.xlsx"

    filepath = os.path.join(EXCEL_OUTPUT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mutual Funds NAV"

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title_font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    subtitle_font = Font(name="Calibri", italic=True, size=10, color="808080")

    data_font = Font(name="Calibri", size=11)
    nav_font = Font(name="Calibri", size=11, bold=True, color="006600")
    category_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # --- Title ---
    ws.merge_cells("A1:I1")
    ws["A1"] = "Pakistan Mutual Funds - Daily NAV Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:I3")
    ws["A3"] = f"Source: www.mufap.com.pk | Total Funds: {len(df)}"
    ws["A3"].font = subtitle_font
    ws["A3"].alignment = Alignment(horizontal="center")

    # --- Headers (row 5) ---
    display_columns = {
        "fund_category": "Fund Category",
        "fund_name": "Fund Name",
        "inception_date": "Inception Date",
        "offer_price": "Offer Price (PKR)",
        "repurchase_price": "Repurchase Price (PKR)",
        "nav": "NAV (PKR)",
        "date_updated": "Validity Date",
        "trustee": "Trustee",
        "scrape_timestamp": "Scraped At",
    }

    available_cols = [col for col in display_columns if col in df.columns]
    headers = [display_columns[col] for col in available_cols]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Data rows ---
    prev_category = None
    for row_idx, (_, record) in enumerate(df.iterrows(), 6):
        for col_idx, col_name in enumerate(available_cols, 1):
            value = record.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Highlight category changes
            current_category = record.get("fund_category", "")
            if current_category != prev_category:
                cell.fill = category_fill

            # Format price/NAV columns
            if col_name in ("nav", "offer_price", "repurchase_price"):
                cell.font = nav_font
                cell.number_format = "#,##0.0000"
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Center dates
            if col_name in ("date_updated", "scrape_timestamp", "inception_date"):
                cell.alignment = Alignment(horizontal="center", vertical="center")

        prev_category = record.get("fund_category", "")

    # --- Column widths ---
    column_widths = {1: 30, 2: 45, 3: 16, 4: 18, 5: 20, 6: 18, 7: 16, 8: 14, 9: 22}
    for col, width in column_widths.items():
        if col <= len(headers):
            ws.column_dimensions[ws.cell(row=5, column=col).column_letter].width = width

    # --- Freeze panes ---
    ws.freeze_panes = "A6"

    # --- Auto-filter ---
    last_col_letter = ws.cell(row=5, column=len(headers)).column_letter
    ws.auto_filter.ref = f"A5:{last_col_letter}{ws.max_row}"

    # --- Summary sheet ---
    if "fund_category" in df.columns:
        ws_summary = wb.create_sheet("Summary")
        ws_summary["A1"] = "Category Summary"
        ws_summary["A1"].font = title_font

        summary = df.groupby("fund_category").agg(
            total_funds=("fund_name", "count"),
            avg_nav=("nav", "mean"),
            min_nav=("nav", "min"),
            max_nav=("nav", "max"),
        ).reset_index()

        summary_headers = ["Fund Category", "Total Funds", "Average NAV", "Min NAV", "Max NAV"]
        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for row_idx, (_, row) in enumerate(summary.iterrows(), 4):
            ws_summary.cell(row=row_idx, column=1, value=row["fund_category"]).border = thin_border
            ws_summary.cell(row=row_idx, column=2, value=row["total_funds"]).border = thin_border
            ws_summary.cell(row=row_idx, column=3, value=round(row["avg_nav"], 4)).border = thin_border
            ws_summary.cell(row=row_idx, column=4, value=round(row["min_nav"], 4)).border = thin_border
            ws_summary.cell(row=row_idx, column=5, value=round(row["max_nav"], 4)).border = thin_border

        for col in range(1, 6):
            ws_summary.column_dimensions[ws_summary.cell(row=3, column=col).column_letter].width = 25

    wb.save(filepath)
    logger.info(f"Excel file saved to: {filepath}")
    return filepath
