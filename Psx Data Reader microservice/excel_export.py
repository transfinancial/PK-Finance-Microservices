"""
Excel Export Module for PSX Data
=================================
Saves PSX stock market data to formatted Excel files.
"""

import os
import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, Reference

from config import EXCEL_OUTPUT_DIR, now_utc5

logger = logging.getLogger(__name__)


def save_stocks_to_excel(df: pd.DataFrame, filename: str = None) -> str:
    """
    Save stock market watch data to a formatted Excel file.

    Args:
        df: DataFrame with stock data
        filename: Optional filename

    Returns:
        Path to saved file
    """
    os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)

    if filename is None:
        timestamp = now_utc5().strftime("%Y%m%d_%H%M%S")
        filename = f"psx_market_data_{timestamp}.xlsx"

    filepath = os.path.join(EXCEL_OUTPUT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Market Watch"

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title_font = Font(name="Calibri", bold=True, size=16, color="1B5E20")
    subtitle_font = Font(name="Calibri", italic=True, size=10, color="808080")

    data_font = Font(name="Calibri", size=10)
    positive_font = Font(name="Calibri", size=10, color="006600", bold=True)
    negative_font = Font(name="Calibri", size=10, color="CC0000", bold=True)
    symbol_font = Font(name="Calibri", size=10, bold=True, color="1B5E20")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    # --- Title ---
    display_cols = ["symbol", "sector_code", "ldcp", "open", "high", "low",
                    "current", "change", "change_pct", "volume", "date"]
    available_cols = [c for c in display_cols if c in df.columns]
    num_cols = len(available_cols)

    last_col_letter = chr(64 + min(num_cols, 26))
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "Pakistan Stock Exchange (PSX) - Market Watch"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Generated: {now_utc5().strftime('%B %d, %Y %I:%M %p')} (UTC+5) | Total Stocks: {len(df)}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # -- Summary stats --
    if "change" in df.columns:
        gainers = len(df[df["change"] > 0])
        losers = len(df[df["change"] < 0])
        unchanged = len(df[df["change"] == 0])

        ws.merge_cells(f"A3:{last_col_letter}3")
        ws["A3"] = f"Gainers: {gainers} | Losers: {losers} | Unchanged: {unchanged}"
        ws["A3"].font = Font(name="Calibri", size=10, bold=True)
        ws["A3"].alignment = Alignment(horizontal="center")

    # --- Headers (row 5) ---
    col_display = {
        "symbol": "Symbol",
        "sector_code": "Sector",
        "ldcp": "LDCP",
        "open": "Open",
        "high": "High",
        "low": "Low",
        "current": "Current",
        "change": "Change",
        "change_pct": "Change %",
        "volume": "Volume",
        "date": "Date",
    }

    headers = [col_display.get(c, c) for c in available_cols]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data rows ---
    for row_idx, (_, record) in enumerate(df.iterrows(), 6):
        for col_idx, col_name in enumerate(available_cols, 1):
            value = record.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = data_font

            # Alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = alt_fill

            # Symbol styling
            if col_name == "symbol":
                cell.font = symbol_font

            # Numeric formatting
            if col_name in ("ldcp", "open", "high", "low", "current"):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

            if col_name == "change":
                change_val = record.get("change", 0)
                if change_val and change_val > 0:
                    cell.font = positive_font
                elif change_val and change_val < 0:
                    cell.font = negative_font
                cell.number_format = "+#,##0.00;-#,##0.00;0.00"
                cell.alignment = Alignment(horizontal="right")

            if col_name == "change_pct":
                cell.number_format = "+#,##0.00%;-#,##0.00%;0.00%"
                change_val = record.get("change", 0)
                if change_val and change_val > 0:
                    cell.font = positive_font
                elif change_val and change_val < 0:
                    cell.font = negative_font
                cell.alignment = Alignment(horizontal="right")

            if col_name == "volume":
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

            if col_name == "date":
                cell.alignment = Alignment(horizontal="center")

    # --- Column widths ---
    width_map = {
        "symbol": 14, "sector_code": 10, "ldcp": 12, "open": 12,
        "high": 12, "low": 12, "current": 12, "change": 12,
        "change_pct": 12, "volume": 16, "date": 14,
    }
    for col_idx, col_name in enumerate(available_cols, 1):
        col_letter = chr(64 + col_idx)
        ws.column_dimensions[col_letter].width = width_map.get(col_name, 14)

    # --- Freeze panes ---
    ws.freeze_panes = "A6"

    # --- Auto-filter ---
    ws.auto_filter.ref = f"A5:{chr(64 + num_cols)}{ws.max_row}"

    # --- Top Gainers / Losers sheets ---
    if "change_pct" in df.columns and len(df) > 0:
        _add_performers_sheet(wb, df, "Top Gainers", ascending=False)
        _add_performers_sheet(wb, df, "Top Losers", ascending=True)

    # --- Volume Leaders sheet ---
    if "volume" in df.columns and len(df) > 0:
        _add_volume_leaders_sheet(wb, df)

    wb.save(filepath)
    logger.info(f"PSX Excel saved: {filepath}")
    return filepath


def _add_performers_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str, ascending: bool):
    """Add a top gainers or losers sheet."""
    ws = wb.create_sheet(sheet_name)

    title_font = Font(name="Calibri", bold=True, size=14,
                      color="006600" if "Gainer" in sheet_name else "CC0000")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="006600" if "Gainer" in sheet_name else "CC0000",
        end_color="006600" if "Gainer" in sheet_name else "CC0000",
        fill_type="solid",
    )
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    sorted_df = df.sort_values("change_pct", ascending=ascending).head(20)

    ws["A1"] = sheet_name
    ws["A1"].font = title_font

    headers = ["Symbol", "Current", "Change", "Change %", "Volume"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(sorted_df.iterrows(), 4):
        ws.cell(row=row_idx, column=1, value=row.get("symbol", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=row.get("current", 0)).border = thin_border
        ws.cell(row=row_idx, column=3, value=row.get("change", 0)).border = thin_border
        ws.cell(row=row_idx, column=4, value=row.get("change_pct", 0)).border = thin_border
        ws.cell(row=row_idx, column=5, value=row.get("volume", 0)).border = thin_border

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 16


def _add_volume_leaders_sheet(wb: Workbook, df: pd.DataFrame):
    """Add volume leaders sheet."""
    ws = wb.create_sheet("Volume Leaders")

    title_font = Font(name="Calibri", bold=True, size=14, color="1565C0")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    sorted_df = df.sort_values("volume", ascending=False).head(20)

    ws["A1"] = "Volume Leaders"
    ws["A1"].font = title_font

    headers = ["Symbol", "Current", "Change %", "Volume"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(sorted_df.iterrows(), 4):
        ws.cell(row=row_idx, column=1, value=row.get("symbol", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=row.get("current", 0)).border = thin_border
        ws.cell(row=row_idx, column=3, value=row.get("change_pct", 0)).border = thin_border
        ws.cell(row=row_idx, column=4, value=row.get("volume", 0)).border = thin_border

    for col in range(1, 5):
        ws.column_dimensions[chr(64 + col)].width = 18
